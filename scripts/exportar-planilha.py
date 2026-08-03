#!/usr/bin/env python3
"""
exportar-planilha.py — regenera a planilha .xlsx a partir de js/dados.js

POR QUE ESTE SCRIPT EXISTE
--------------------------
A planilha era, originalmente, a fonte dos dados: ela alimentava
data/*.json via gerar_dados.py. Esse fluxo foi invertido — js/dados.js
passou a ser a fonte única — e a planilha ficou para trás, parada em 97
registros enquanto o site já tinha 176.

Agora ela é um ARTEFATO GERADO, como data/*.json: continua útil para
consultar, filtrar e imprimir, mas não se edita mais à mão. Editar aqui
e rodar o obsoleto gerar_dados.py destruiria o banco.

Preserva a estrutura da planilha original: título, subtítulo, a mesma
ordem das 13 colunas e a aba "Legenda e Fontes". Os campos que só
existem no banco (id, sítio, coordenadas, bacia, DOI) entram como
colunas ADICIONAIS ao final, sem deslocar as originais.

Uso:
    python3 scripts/exportar-planilha.py
"""
import json, re, pathlib
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RAIZ = pathlib.Path(__file__).resolve().parent.parent
FONTE = RAIZ / "js" / "dados.js"
SAIDA = RAIZ / "fosseis_santa_catarina_enriquecido.xlsx"

FONTE_TXT = "Arial"
PETROL = "1F4E5F"
PEDRA = "DDE1D3"

# (cabeçalho, chave em DB_FOSSEIS, largura)
COLUNAS = [
    ("Período Geológico",              "periodo",          26),
    ("Formação / Grupo",               "formacao",         34),
    ("Táxon / Material",               "taxon",            40),
    ("Categoria",                      "categoria",        26),
    ("Nº de Amostra / Catálogo",       "numero_catalogo",  22),
    ("Local de Coleta (Município)",    "local_coleta",     38),
    ("Local de Armazenamento",         "armazenamento",    34),
    ("Descritor(es) / Referência",     "descritor",        40),
    ("Unidade de Pesquisa",            "unidade_pesquisa", 28),
    ("Idade Estimada (Ma)",            "idade_ma",         24),
    ("Observações",                    "observacoes",      60),
    ("Fonte / URL de Verificação",     "fontes",           44),
    ("Citação Científica (ABNT)",      "citacao_abnt",     50),
    # colunas adicionais — só existem no banco
    ("Nº do registro",                 "id",               13),
    ("Sítio",                          "site",             28),
    ("Bacia / Unidade",                "bacia",            30),
    ("Latitude",                       "lat",              11),
    ("Longitude",                      "lon",              11),
    ("DOI",                            "doi",              24),
]

def carregar(nome):
    txt = FONTE.read_text(encoding="utf-8")
    m = re.search(r"^const %s = (\[.*?\]);\n" % nome, txt, re.S | re.M)
    return json.loads(m.group(1))

def valor(reg, chave):
    v = reg.get(chave)
    if v is None:
        return ""
    if isinstance(v, list):
        return "\n".join(str(x) for x in v)
    return v

def main():
    foss = sorted(carregar("DB_FOSSEIS"), key=lambda d: (d["periodo_ordem"], d["id"]))

    # preserva a aba de legenda da planilha existente, se houver
    legenda = []
    if SAIDA.exists():
        try:
            antigo = openpyxl.load_workbook(SAIDA)
            if "Legenda e Fontes" in antigo.sheetnames:
                la = antigo["Legenda e Fontes"]
                legenda = [[c.value for c in linha] for linha in la.iter_rows()]
        except Exception:
            pass

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catálogo de Fósseis SC"

    fina = Side(style="thin", color="BBBBBB")
    borda = Border(left=fina, right=fina, top=fina, bottom=fina)

    ws["A1"] = "BANCO DE DADOS PALEONTOLÓGICO — ESTADO DE SANTA CATARINA (Brasil)"
    ws["A1"].font = Font(name=FONTE_TXT, size=14, bold=True, color=PETROL)
    ws["A2"] = (f"{len(foss)} registros · GERADO automaticamente a partir de js/dados.js "
                f"por scripts/exportar-planilha.py — NÃO editar à mão: as alterações se perdem "
                f"na próxima geração. Compilado por Brenno Alef Benk.")
    ws["A2"].font = Font(name=FONTE_TXT, size=9, italic=True, color="777777")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUNAS))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLUNAS))

    for i, (cab, _, larg) in enumerate(COLUNAS, start=1):
        cel = ws.cell(row=3, column=i, value=cab)
        cel.font = Font(name=FONTE_TXT, size=10, bold=True, color="FFFFFF")
        cel.fill = PatternFill("solid", fgColor=PETROL)
        cel.alignment = Alignment(vertical="center", wrap_text=True)
        cel.border = borda
        ws.column_dimensions[get_column_letter(i)].width = larg
    ws.row_dimensions[3].height = 30

    for linha, reg in enumerate(foss, start=4):
        for i, (_, chave, _) in enumerate(COLUNAS, start=1):
            cel = ws.cell(row=linha, column=i, value=valor(reg, chave))
            cel.font = Font(name=FONTE_TXT, size=9)
            cel.alignment = Alignment(vertical="top", wrap_text=chave in
                ("observacoes", "citacao_abnt", "fontes", "descritor", "local_coleta", "taxon"))
            cel.border = borda
        if linha % 2 == 0:
            for i in range(1, len(COLUNAS) + 1):
                ws.cell(row=linha, column=i).fill = PatternFill("solid", fgColor="F6F7F0")

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(COLUNAS))}{len(foss)+3}"

    # aba de legenda
    wl = wb.create_sheet("Legenda e Fontes")
    if legenda:
        for r, linha in enumerate(legenda, start=1):
            for cidx, v in enumerate(linha, start=1):
                cel = wl.cell(row=r, column=cidx, value=v)
                cel.font = Font(name=FONTE_TXT, size=10, bold=(r == 1))
                cel.alignment = Alignment(vertical="top", wrap_text=True)
    else:
        wl["A1"], wl["B1"] = "Campo", "Descrição / Critério"
        wl["A1"].font = wl["B1"].font = Font(name=FONTE_TXT, size=10, bold=True)
    wl.column_dimensions["A"].width = 30
    wl.column_dimensions["B"].width = 100

    prox = wl.max_row + 2
    notas = [
        ("PROCEDÊNCIA DOS DADOS", ""),
        ("Fonte única", "js/dados.js. Esta planilha e data/*.json são gerados a partir dele."),
        ("Regenerar", "python3 scripts/exportar-planilha.py"),
        ("Validar", "python3 scripts/validar.py"),
        ("gerar_dados.py", "OBSOLETO. Ia da planilha para data/*.json; rodá-lo sobrescreveria o banco com dados antigos."),
        ("Citação ABNT", "Preenchida nos 97 registros originais. Os acrescentados depois ainda não têm — o campo fica vazio."),
        ("Escopo", "Apenas material COLETADO em Santa Catarina. Guarda em outro estado não descaracteriza o registro."),
        ("Viés amostral", "A densidade reflete esforço de publicação, não riqueza fossilífera."),
    ]
    for r, (a, b) in enumerate(notas, start=prox):
        wl.cell(row=r, column=1, value=a).font = Font(name=FONTE_TXT, size=10, bold=not b)
        wl.cell(row=r, column=2, value=b).font = Font(name=FONTE_TXT, size=10)
        wl.cell(row=r, column=2).alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(SAIDA)
    sem_cit = sum(1 for d in foss if not d.get("citacao_abnt"))
    print(f"Planilha gerada: {SAIDA.name}")
    print(f"  {len(foss)} registros · {len(COLUNAS)} colunas")
    print(f"  {len(foss)-sem_cit} com citação ABNT · {sem_cit} sem (campo vazio)")

if __name__ == "__main__":
    main()
